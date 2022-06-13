from typing import Tuple
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from os import listdir, remove
from os.path import exists
import warnings
warnings.filterwarnings('ignore')


def choose_dir():
    exams = listdir('./试卷')
    for index, dir in enumerate(exams):
        print(f'{index}) {dir}')
    choice = input('enter the index of the exam: ')
    return exams[int(choice)]


def delete_file(path):
    if exists(path):
        remove(path)


def handle_scores_file(path):
    wb = load_workbook(path)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.value is not None:
                cell.value = float(cell.value)
    wb.save(path)


def handle_scores_file(path):
    wb = load_workbook(path)
    ws = wb.active
    ws = delete_multi(ws, 'row', (1, 2))

    ty_start = -1
    max_col = ws.max_column
    for i in range(1, max_col+1):
        # 组合单元格取值错误
        print(ws.cell(row=3, column=i).value)
        if str(ws.cell(row=3, column=i).value)[0] == '二':
            ty_start = i
            break
    delete_col = ((1, 1), (4, 2), (9, 2), (ty_start, max_col))
    ws = delete_multi(ws, 'col', *delete_col)

    wb.save(path)
    print(path, 'success')


def handle_analysis_file(path):
    pass


def delete_multi(ws: Worksheet, rc: str, *areas: Tuple[int, int]):
    if rc == 'col':
        for area in areas:
            ws.delete_cols(area[0], area[1])
    if rc == 'row':
        for area in areas:
            ws.delete_rows(area[0], area[1])
    return ws


if __name__ == '__main__':
    dir = f'./试卷/{choose_dir()}/成绩'
    classes = listdir(dir)  # 获取班级列表
    for c in classes:
        if c[0] == '高':
            classdir = f'{dir}/{c}'
            files = listdir(classdir)  # 获取班级下的文件列表
            for file in files:
                if '客观题填涂表' in file or '技术分段统计' in file:
                    delete_file(f'{classdir}/{file}')
                if '技术成绩' in file:
                    handle_scores_file(f'{classdir}/{file}')
                if '试卷分析' in file:
                    handle_analysis_file(f'{classdir}/{file}')
